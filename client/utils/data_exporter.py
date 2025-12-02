"""
数据导出工具
"""
import sys
from pathlib import Path
import json
import csv
from datetime import datetime
from typing import List, Optional
from loguru import logger

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from client.models import SessionLocal, Regulation


class DataExporter:
    """数据导出工具"""

    def __init__(self):
        self.db = SessionLocal()

    def __del__(self):
        if hasattr(self, 'db'):
            self.db.close()

    def export_to_json(self, output_path: str, regulations: Optional[List[Regulation]] = None) -> tuple[bool, str]:
        """导出为JSON格式"""
        try:
            if regulations is None:
                regulations = self.db.query(Regulation).all()

            data = {
                "export_time": datetime.now().isoformat(),
                "total_count": len(regulations),
                "regulations": []
            }

            for reg in regulations:
                reg_data = {
                    "code": reg.code,
                    "name": reg.name,
                    "country": reg.country,
                    "category": reg.category,
                    "description": reg.description,
                    "status": reg.status.value,
                    "version": reg.version,
                    "tags": [tag.name for tag in reg.tags],
                    "created_at": reg.created_at.isoformat() if reg.created_at else None,
                }
                data["regulations"].append(reg_data)

            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            logger.info(f"成功导出 {len(regulations)} 条法规到 {output_path}")
            return True, f"成功导出 {len(regulations)} 条法规"

        except Exception as e:
            logger.error(f"导出JSON失败: {e}")
            return False, f"导出失败: {str(e)}"

    def export_to_csv(self, output_path: str, regulations: Optional[List[Regulation]] = None) -> tuple[bool, str]:
        """导出为CSV格式"""
        try:
            if regulations is None:
                regulations = self.db.query(Regulation).all()

            fieldnames = ['法规编号', '法规名称', '国家/地区', '分类', '状态', '版本', '标签', '描述']

            with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()

                for reg in regulations:
                    writer.writerow({
                        '法规编号': reg.code,
                        '法规名称': reg.name,
                        '国家/地区': reg.country or '',
                        '分类': reg.category or '',
                        '状态': reg.status.value,
                        '版本': reg.version or '',
                        '标签': ', '.join([tag.name for tag in reg.tags]),
                        '描述': reg.description or '',
                    })

            logger.info(f"成功导出 {len(regulations)} 条法规")
            return True, f"成功导出 {len(regulations)} 条法规"

        except Exception as e:
            logger.error(f"导出CSV失败: {e}")
            return False, f"导出失败: {str(e)}"

    def export_to_excel(self, output_path: str, regulations: Optional[List[Regulation]] = None) -> tuple[bool, str]:
        """导出为Excel格式"""
        try:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font
            except ImportError:
                return False, "需要安装openpyxl库: pip install openpyxl"

            if regulations is None:
                regulations = self.db.query(Regulation).all()

            wb = Workbook()
            ws = wb.active
            ws.title = "法规列表"

            headers = ['法规编号', '法规名称', '国家/地区', '分类', '状态', '版本', '标签', '描述']

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)

            for row, reg in enumerate(regulations, 2):
                ws.cell(row=row, column=1, value=reg.code)
                ws.cell(row=row, column=2, value=reg.name)
                ws.cell(row=row, column=3, value=reg.country or '')
                ws.cell(row=row, column=4, value=reg.category or '')
                ws.cell(row=row, column=5, value=reg.status.value)
                ws.cell(row=row, column=6, value=reg.version or '')
                ws.cell(row=row, column=7, value=', '.join([tag.name for tag in reg.tags]))
                ws.cell(row=row, column=8, value=reg.description or '')

            wb.save(output_path)
            logger.info(f"成功导出 {len(regulations)} 条法规")
            return True, f"成功导出 {len(regulations)} 条法规"

        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            return False, f"导出失败: {str(e)}"