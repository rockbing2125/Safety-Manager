"""
数据导入工具
"""
import sys
from pathlib import Path
import json
import csv
from typing import Tuple, List, Dict
from loguru import logger

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from client.models import SessionLocal, Regulation, Tag
from shared.constants import RegulationStatus
from client.services import RegulationService


class DataImporter:
    """数据导入工具"""

    def __init__(self):
        self.db = SessionLocal()
        self.regulation_service = RegulationService()

    def __del__(self):
        if hasattr(self, 'db'):
            self.db.close()

    def import_from_json(self, file_path: str, user_id: int,
                        overwrite: bool = False) -> Tuple[bool, str, Dict]:
        """
        从JSON文件导入法规

        Args:
            file_path: JSON文件路径
            user_id: 用户ID
            overwrite: 如果法规编号已存在，是否覆盖

        Returns:
            (成功, 消息, 统计信息)
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            regulations_data = data.get('regulations', [])
            if not regulations_data:
                return False, "文件中没有找到法规数据", {}

            stats = {
                'total': len(regulations_data),
                'success': 0,
                'skipped': 0,
                'failed': 0,
                'errors': []
            }

            for reg_data in regulations_data:
                try:
                    code = reg_data.get('code')
                    if not code:
                        stats['failed'] += 1
                        stats['errors'].append("缺少法规编号")
                        continue

                    # 检查是否已存在
                    existing = self.db.query(Regulation).filter(
                        Regulation.code == code
                    ).first()

                    if existing and not overwrite:
                        stats['skipped'] += 1
                        continue

                    if existing and overwrite:
                        # 更新现有法规
                        existing.name = reg_data.get('name', existing.name)
                        existing.country = reg_data.get('country')
                        existing.category = reg_data.get('category')
                        existing.description = reg_data.get('description')
                        existing.version = reg_data.get('version')

                        # 更新状态
                        status_value = reg_data.get('status', 'active')
                        try:
                            existing.status = RegulationStatus(status_value)
                        except ValueError:
                            existing.status = RegulationStatus.ACTIVE

                        # 更新标签
                        if 'tags' in reg_data:
                            existing.tags.clear()
                            for tag_name in reg_data.get('tags', []):
                                tag = self.db.query(Tag).filter(
                                    Tag.name == tag_name
                                ).first()
                                if not tag:
                                    tag = Tag(name=tag_name)
                                    self.db.add(tag)
                                existing.tags.append(tag)

                        stats['success'] += 1
                    else:
                        # 创建新法规
                        status_value = reg_data.get('status', 'active')
                        try:
                            status = RegulationStatus(status_value)
                        except ValueError:
                            status = RegulationStatus.ACTIVE

                        regulation = Regulation(
                            code=code,
                            name=reg_data.get('name', ''),
                            country=reg_data.get('country'),
                            category=reg_data.get('category'),
                            description=reg_data.get('description'),
                            status=status,
                            version=reg_data.get('version'),
                            created_by=user_id
                        )

                        # 添加标签
                        for tag_name in reg_data.get('tags', []):
                            tag = self.db.query(Tag).filter(
                                Tag.name == tag_name
                            ).first()
                            if not tag:
                                tag = Tag(name=tag_name)
                                self.db.add(tag)
                            regulation.tags.append(tag)

                        self.db.add(regulation)
                        stats['success'] += 1

                except Exception as e:
                    logger.error(f"导入法规失败: {reg_data.get('code', 'unknown')}, {e}")
                    stats['failed'] += 1
                    stats['errors'].append(f"{reg_data.get('code', 'unknown')}: {str(e)}")

            self.db.commit()

            message = f"导入完成！\n成功: {stats['success']}, 跳过: {stats['skipped']}, 失败: {stats['failed']}"
            logger.info(f"JSON导入完成: {message}")
            return True, message, stats

        except Exception as e:
            self.db.rollback()
            logger.error(f"导入JSON失败: {e}")
            return False, f"导入失败: {str(e)}", {}

    def import_from_excel(self, file_path: str, user_id: int,
                         overwrite: bool = False) -> Tuple[bool, str, Dict]:
        """
        从Excel文件导入法规

        Args:
            file_path: Excel文件路径
            user_id: 用户ID
            overwrite: 如果法规编号已存在，是否覆盖

        Returns:
            (成功, 消息, 统计信息)
        """
        try:
            try:
                from openpyxl import load_workbook
            except ImportError:
                return False, "需要安装openpyxl库: pip install openpyxl", {}

            wb = load_workbook(file_path, read_only=True)
            ws = wb.active

            # 读取表头
            headers = [cell.value for cell in ws[1]]

            # 查找列索引
            col_map = {}
            for idx, header in enumerate(headers):
                if header in ['法规编号', 'code', 'Code', '编号']:
                    col_map['code'] = idx
                elif header in ['法规名称', 'name', 'Name', '名称']:
                    col_map['name'] = idx
                elif header in ['国家/地区', 'country', 'Country', '国家']:
                    col_map['country'] = idx
                elif header in ['分类', 'category', 'Category']:
                    col_map['category'] = idx
                elif header in ['状态', 'status', 'Status']:
                    col_map['status'] = idx
                elif header in ['版本', 'version', 'Version']:
                    col_map['version'] = idx
                elif header in ['标签', 'tags', 'Tags']:
                    col_map['tags'] = idx
                elif header in ['描述', 'description', 'Description']:
                    col_map['description'] = idx

            if 'code' not in col_map or 'name' not in col_map:
                return False, "Excel文件缺少必需的列：法规编号、法规名称", {}

            stats = {
                'total': ws.max_row - 1,
                'success': 0,
                'skipped': 0,
                'failed': 0,
                'errors': []
            }

            # 跳过表头，从第2行开始
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    code = row[col_map['code']]
                    if not code:
                        stats['failed'] += 1
                        stats['errors'].append(f"第{row_idx}行: 缺少法规编号")
                        continue

                    name = row[col_map['name']]
                    if not name:
                        stats['failed'] += 1
                        stats['errors'].append(f"第{row_idx}行: 缺少法规名称")
                        continue

                    # 检查是否已存在
                    existing = self.db.query(Regulation).filter(
                        Regulation.code == str(code)
                    ).first()

                    if existing and not overwrite:
                        stats['skipped'] += 1
                        continue

                    # 获取其他字段
                    country = row[col_map['country']] if 'country' in col_map else None
                    category = row[col_map['category']] if 'category' in col_map else None
                    description = row[col_map['description']] if 'description' in col_map else None
                    version = row[col_map['version']] if 'version' in col_map else None

                    status_value = row[col_map['status']] if 'status' in col_map else 'active'
                    try:
                        status = RegulationStatus(status_value.lower() if status_value else 'active')
                    except (ValueError, AttributeError):
                        status = RegulationStatus.ACTIVE

                    tags_str = row[col_map['tags']] if 'tags' in col_map else ''
                    tag_names = [t.strip() for t in str(tags_str).split(',') if t.strip()] if tags_str else []

                    if existing and overwrite:
                        # 更新现有法规
                        existing.name = str(name)
                        existing.country = str(country) if country else None
                        existing.category = str(category) if category else None
                        existing.description = str(description) if description else None
                        existing.version = str(version) if version else None
                        existing.status = status

                        # 更新标签
                        existing.tags.clear()
                        for tag_name in tag_names:
                            tag = self.db.query(Tag).filter(
                                Tag.name == tag_name
                            ).first()
                            if not tag:
                                tag = Tag(name=tag_name)
                                self.db.add(tag)
                            existing.tags.append(tag)

                        stats['success'] += 1
                    else:
                        # 创建新法规
                        regulation = Regulation(
                            code=str(code),
                            name=str(name),
                            country=str(country) if country else None,
                            category=str(category) if category else None,
                            description=str(description) if description else None,
                            status=status,
                            version=str(version) if version else None,
                            created_by=user_id
                        )

                        # 添加标签
                        for tag_name in tag_names:
                            tag = self.db.query(Tag).filter(
                                Tag.name == tag_name
                            ).first()
                            if not tag:
                                tag = Tag(name=tag_name)
                                self.db.add(tag)
                            regulation.tags.append(tag)

                        self.db.add(regulation)
                        stats['success'] += 1

                except Exception as e:
                    logger.error(f"导入第{row_idx}行失败: {e}")
                    stats['failed'] += 1
                    stats['errors'].append(f"第{row_idx}行: {str(e)}")

            self.db.commit()
            wb.close()

            message = f"导入完成！\n成功: {stats['success']}, 跳过: {stats['skipped']}, 失败: {stats['failed']}"
            logger.info(f"Excel导入完成: {message}")
            return True, message, stats

        except Exception as e:
            self.db.rollback()
            logger.error(f"导入Excel失败: {e}")
            return False, f"导入失败: {str(e)}", {}
